from .MongoClient import MongoClientAsync
from etc.Globals import ProductMapping, Product
from abc import ABC, abstractmethod
import json
from openpyxl import Workbook, load_workbook
from string import ascii_uppercase

class ProductUploader(ABC):
    UPSERT = True
    DEFAULT_QUERY_FILTER = "_id"
    def __init__(self, isuri: str, database: str, collection: str):
        assert isinstance(isuri, str)
        self._database = database
        self._collection = collection
        self._client = MongoClientAsync(isuri)
        self._client.select_collection(database, collection)

    @abstractmethod
    async def push_changes(self, filename: str, query_filter_name: str) -> int:
        pass

    @abstractmethod
    def write_file(self, products: dict, filename: str):
        pass

    async def get_products(self) -> dict:
        ''' Get all products in a dictionary, each tabled
        by _id.'''
        product_json = {}
        cursor = await self._client.find({})

        async for item in cursor:
            assert isinstance(item, dict)
            product_json[item["_id"]] = item

        return product_json

    async def _push_dict(self, data_dict: dict, query_filter_name: str) -> int:
        ''' Replace all data in the previously selected
        database and collection with the given dictionary.

        data_dict: dictionary containing all data to replace,
                    indexed by _id.

        query_filter_name: the filter to apply when searching for the product
                            to replace.

        Returns: the amount of products replaced.'''
        assert isinstance(data_dict, dict)
        assert isinstance(query_filter_name, str)
        doc_pairs = []

        for _, item in data_dict.items():
            assert isinstance(item, dict), f"{item} is not a dict!"
            query_filter = {query_filter_name: item[query_filter_name]}
            data_to_push = item
            doc_pairs.append((query_filter, data_to_push, self.UPSERT))

        await self._client.bulk_replace(doc_pairs)
        return len(doc_pairs)
    
class SpreadsheetProductUploader(ProductUploader):
    ''' Allow client to write changes to a spreadsheet
    and have those changes pushed to the item server.

    This class is a containment of JSONProductUploader, as
    it is necessary to convert all spreadsheet data
    into json before uploading.
    '''
    DEFAULT_OUTPUT_FILENAME = "products.xlsx"
    async def push_changes(self, filename: str, query_filter_name: str=ProductUploader.DEFAULT_QUERY_FILTER):
        ''' Push changes from the supplied filename
        onto database.collection.

        filename: name of the spreadsheet file

        query_filter_name: filter to apply when searching
                            for products to replace
        
        returns: the amount of products replaced.
        '''
        workbook = load_workbook(filename=filename)
        products_dict = self._workbook_to_dict(workbook)
        return await self._push_dict(products_dict, query_filter_name)

    async def write_file(self, products: dict, filename: str):
        workbook = await self._json_to_workbook(products)
        workbook.save(filename)

    async def _json_to_workbook(self, products: dict) -> Workbook:
        ''' Convert all dictionaries from json to
        a workbook.

        Returns a workbook with a formatted sheet'''
        workbook = Workbook()
        sheet = workbook.active

        # Initialize the field row
        for prod in ProductMapping:
            assert isinstance(prod.value, int)
            assert isinstance(prod.name, str)
            sheet[f"{self._int_to_alpha(prod.value)}1"] = prod.name

        # Append new products to the sheet
        curr_row = 2

        for _, product in products.items():
            for key, val in self._denest_dict(product).items():
                if isinstance(val, str):
                    column = ProductMapping.get(key).value
                    sheet[f"{self._int_to_alpha(column)}{curr_row}"] = val

            curr_row += 1
        return workbook

    def _workbook_to_dict(self, workbook: Workbook) -> dict:
        assert isinstance(workbook, Workbook)
        ''' Convert the workbook's sheet of products
        into the corresponding product dictionary.'''
        sheet = workbook.active

        # Verify that the header row is correctly formatted.
        if self._verify_header_row(tuple([cell.value for cell in sheet[1]])):
            products_dict = {}

            # Get all products from each row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                prod = Product()
                
                for i, val in enumerate(row):
                    prod.__dict__[ProductMapping.get(i + 1).name] = val if val else ""
        
                products_dict[prod._id] = prod.as_dict()
            return products_dict
        return None

    def _verify_header_row(self, row: tuple):
        if len(row) != len(ProductMapping):
            return False

        for i, field in enumerate(row):
            try:
                if i != ProductMapping.get(field).value - 1:
                    return False

            except KeyError:
                return False
        return True

    def _int_to_alpha(self, val: int) -> str:
        ''' Convert the integer value to its respective
        alphanumeric string. E.g.:
        1 -> "A"
        2 -> "B"
        26 -> "Z"
        27 -> "AA"
        28 -> ""AB"
        53 -> "BA"

        Treat the strings as if they were a base-26 system.'''
        assert isinstance(val, int)
        
        # TODO: provide functionality for int vals past 26
        if val >= 1:
            return ascii_uppercase[val - 1]

        raise ValueError(f"{val} isn't >= 1")

    def _denest_dict(self, dictionary: dict) -> dict:
        ''' Denest all dictionaries within the given dictionary and
            return a dictionary with all fields of non-dictionary
            class.'''
        assert isinstance(dictionary, dict)
        dicts_found = []
        denested_dict = {}

        # Iterate through each field to check if
        # a nested dictionary is present.
        for key, item in dictionary.items():
            if isinstance(item, dict):
                dicts_found.append(item)

            else:
                denested_dict[key] = item

        # Base case: the current dictionary doesn't contain
        # any nested dictionaries. Denested_dict is already
        # populated with the same fields as the original dictionary.
        if len(dicts_found) == 0:
            return denested_dict

        # Recursive case: the current dictionary contains >= 1
        # dictionary items.
        else:
            [denested_dict.update(self._denest_dict(d)) for d in dicts_found]
            return denested_dict
            
class JSONProductUploader(ProductUploader):
    ''' Allow client to write changes to a json file
    and have those changes pushed to the item server.

    Returns the amount of items pushed.'''

    DEFAULT_OUTPUT_FILENAME = "products.json"
    DEFAULT_INDENT = 4
    def __init__(self, isuri: str, database: str, collection: str):
        ProductUploader.__init__(self, isuri, database, collection)

    async def push_changes(self, filename: str, query_filter_name: str=ProductUploader.DEFAULT_QUERY_FILTER) -> int:
        assert isinstance(filename, str)
        assert isinstance(query_filter_name, str)
        assert query_filter_name != ""

        with open(filename, "r") as rfile:
            data_dict = json.load(rfile)
            products_replaced = await self._push_dict(data_dict, query_filter_name)
            return products_replaced

    async def write_file(self, products: dict, filename: str):
        assert isinstance(products, dict)
        assert isinstance(filename, str)
        with open(filename, "w") as wfile:
            json.dump(products, wfile, indent=self.DEFAULT_INDENT)

